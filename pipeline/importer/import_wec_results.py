"""Import WEC ward-by-ward canvass spreadsheets into election_history.

Usage: python -m importer.import_wec_results <xlsx_dir> <sqlite_path>

Each contest appears as a title row, a party header containing 'Total Votes
Cast', a candidate row, then ward rows whose votes are summed per candidate.
"""

from __future__ import annotations

import argparse
import re
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

CONTEST_RE = re.compile(
    r"^(STATE SENATOR|REPRESENTATIVE TO THE ASSEMBLY)\s+DISTRICT\s+(\d+)", re.I
)
STATEWIDE_RE = re.compile(
    r"^(GOVERNOR / LIEUTENANT GOVERNOR|ATTORNEY GENERAL"
    r"|SECRETARY OF STATE|STATE TREASURER)$",
    re.I,
)
YEAR_RE = re.compile(r"^(\d{4}) General Election", re.I)


def _ticket(candidate: str) -> str:
    """'Tony Evers \\n Sara Rodriguez' -> 'Tony Evers / Sara Rodriguez'."""
    parts = [" ".join(p.split()) for p in str(candidate).split("\n")]
    return " / ".join(p for p in parts if p)


SUBTOTAL_RE = re.compile(r"totals?:\s*$", re.I)


def _walk_contests(path: Path, contest_re: re.Pattern):
    """Yield (year, match, parties, candidates, sums, total_cast, counties)
    per contest.

    Columns anchor absolutely on the 'Total Votes Cast' header (canvass
    rows are ragged, so length arithmetic misaligns). Subtotal rows are
    excluded from the ward sums — the report interleaves them, which
    silently multiplies every total — but the 'County Totals:' rows are
    themselves certified county aggregates, so they're captured as
    (county, per-candidate votes) instead of discarded."""
    wb = load_workbook(path, read_only=True)
    year: int | None = None
    match: re.Match | None = None
    base: int | None = None
    parties: list[str | None] = []
    candidates: list[str] = []
    totals: list[int] = []
    total_cast = 0
    county: str | None = None
    counties: list[tuple[str, list[int]]] = []

    def snapshot():
        return (year, match, list(parties), list(candidates), list(totals),
                total_cast, list(counties))

    def reset(m):
        nonlocal match, base, parties, candidates, totals, total_cast, county, counties
        match, base, parties, candidates = m, None, [], []
        totals, total_cast, county, counties = [], 0, None, []

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            texts = ["" if c is None else str(c).strip() for c in cells]
            first = texts[0] if texts else ""
            if year is None:
                m = YEAR_RE.match(first)
                if m:
                    year = int(m.group(1))
            m = contest_re.match(first)
            if m:
                if match and candidates:
                    yield snapshot()
                reset(m)
                continue
            if match is None:
                continue
            if "Total Votes Cast" in texts:
                base = texts.index("Total Votes Cast") + 1
                parties, candidates, totals = [t or None for t in texts[base:]], [], []
                continue
            if base is not None and not candidates and any(texts[base:]):
                candidates = texts[base:]
                while candidates and not candidates[-1]:
                    candidates.pop()
                totals = [0] * len(candidates)
                parties = (parties + [None] * len(candidates))[: len(candidates)]
                continue
            if candidates:
                if any(SUBTOTAL_RE.search(t) for t in texts[:2] if t):
                    # the county aggregate row: capture once, never sum
                    if texts[1] == "County Totals:" and county:
                        counties.append((
                            county,
                            [int(cells[base + i]) if base + i < len(cells)
                             and isinstance(cells[base + i], (int, float)) else 0
                             for i in range(len(candidates))],
                        ))
                    continue
                if first:
                    county = " ".join(first.split()).title()
                cast = cells[base - 1] if base - 1 < len(cells) else None
                if isinstance(cast, (int, float)):
                    total_cast += int(cast)
                for i in range(len(candidates)):
                    v = cells[base + i] if base + i < len(cells) else None
                    if isinstance(v, (int, float)):
                        totals[i] += int(v)
    if match and candidates:
        yield snapshot()
    wb.close()


def parse_workbook(path: Path) -> list[tuple[int, str, int, str, str | None, int, int]]:
    rows_out: list[tuple[int, str, int, str, str | None, int, int]] = []
    for year, m, parties, candidates, totals, cast, _ in _walk_contests(path, CONTEST_RE):
        if year is None:
            raise RuntimeError(f"{path.name}: contest found before year header")
        chamber = "upper" if m.group(1).upper().startswith("STATE SEN") else "lower"
        district = int(m.group(2))
        for i, candidate in enumerate(candidates):
            if not candidate or candidate.upper() == "SCATTERING":
                continue
            rows_out.append(
                (year, chamber, district, " ".join(str(candidate).split()),
                 parties[i], totals[i], cast)
            )
    return rows_out


def parse_statewide(path: Path):
    """Returns (candidate rows incl. total_cast, county rows)."""
    rows_out: list[tuple[int, str, str, str | None, int, int]] = []
    county_rows: list[tuple[int, str, str, str, str | None, int]] = []
    for year, m, parties, candidates, totals, cast, counties in _walk_contests(
        path, STATEWIDE_RE
    ):
        if year is None:
            raise RuntimeError(f"{path.name}: contest found before year header")
        office = m.group(1).upper()
        for i, candidate in enumerate(candidates):
            if not candidate or candidate.upper() == "SCATTERING":
                continue
            rows_out.append(
                (year, office, _ticket(candidate), parties[i], totals[i], cast)
            )
            for county, values in counties:
                county_rows.append(
                    (year, office, county, _ticket(candidate), parties[i], values[i])
                )
    return rows_out, county_rows


def run(xlsx_dir: Path, db_path: Path) -> int:
    all_rows: list[tuple] = []
    statewide_rows: list[tuple] = []
    county_rows: list[tuple] = []
    for path in sorted(xlsx_dir.glob("*.xlsx")):
        rows = parse_workbook(path)
        sw, counties = parse_statewide(path)
        print(f"{path.name}: {len(rows)} candidate rows"
              + (f", {len(sw)} statewide" if sw else ""))
        all_rows.extend(rows)
        statewide_rows.extend(sw)
        county_rows.extend(counties)
    if not all_rows:
        raise RuntimeError(f"no legislative contests parsed from {xlsx_dir}")
    # a general election covers all 99 Assembly districts
    for year in {r[0] for r in all_rows}:
        assembly = {r[2] for r in all_rows if r[0] == year and r[1] == "lower"}
        if len(assembly) < 95:
            raise RuntimeError(
                f"{year}: only {len(assembly)} Assembly districts parsed — format drift?"
            )
    # the official Total Votes Cast can never be below the candidate sum
    for key in {(r[0], r[1], r[2]) for r in all_rows}:
        contest = [r for r in all_rows if (r[0], r[1], r[2]) == key]
        if contest[0][6] < sum(r[5] for r in contest):
            raise RuntimeError(f"{key}: total cast below candidate sum — bad parse")
    # every statewide contest is a two-party-plus race with 7-figure turnout,
    # its 72 county aggregates must sum exactly to the statewide totals
    for year, office in {(r[0], r[1]) for r in statewide_rows}:
        contest = [r for r in statewide_rows if r[0] == year and r[1] == office]
        total = sum(r[4] for r in contest)
        if len(contest) < 2 or total < 500_000 or contest[0][5] < total:
            raise RuntimeError(
                f"{year} {office}: {len(contest)} candidates, {total} votes"
                " — format drift?"
            )
        counties = {r[2] for r in county_rows if r[0] == year and r[1] == office}
        if len(counties) != 72:
            raise RuntimeError(
                f"{year} {office}: {len(counties)} counties parsed, expected 72"
            )
        for r in contest:
            county_sum = sum(
                c[5] for c in county_rows
                if c[0] == year and c[1] == office and c[3] == r[2]
            )
            if county_sum != r[4]:
                raise RuntimeError(
                    f"{year} {office} {r[2]}: county sum {county_sum}"
                    f" != statewide {r[4]}"
                )

    conn = sqlite3.connect(db_path)
    with conn:
        conn.execute("DELETE FROM election_history")
        conn.executemany(
            "INSERT INTO election_history (year, chamber, district, candidate,"
            " party, votes, total_cast) VALUES (?, ?, ?, ?, ?, ?, ?)",
            all_rows,
        )
        conn.execute("DELETE FROM statewide_history")
        conn.executemany(
            "INSERT INTO statewide_history (year, office, candidate, party, votes,"
            " total_cast) VALUES (?, ?, ?, ?, ?, ?)",
            statewide_rows,
        )
        conn.execute("DELETE FROM statewide_county_results")
        conn.executemany(
            "INSERT INTO statewide_county_results (year, office, county, candidate,"
            " party, votes) VALUES (?, ?, ?, ?, ?, ?)",
            county_rows,
        )
    seats = conn.execute(
        "SELECT COUNT(DISTINCT chamber || district || year) FROM election_history"
    ).fetchone()[0]
    conn.close()
    print(f"election_history: {len(all_rows)} rows across {seats} contests")
    contests = len({(r[0], r[1]) for r in statewide_rows})
    print(f"statewide_history: {len(statewide_rows)} rows across {contests} contests;"
          f" {len(county_rows)} county result rows")
    return 0


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx_dir", type=Path)
    parser.add_argument("db_path", type=Path)
    ns = parser.parse_args(argv)
    return run(ns.xlsx_dir, ns.db_path)


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
